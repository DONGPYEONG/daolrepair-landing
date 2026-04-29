# -*- coding: utf-8 -*-
"""
다올리페어 네이버 파워링크 — 키워드/카피/예산 Excel 생성기
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'naver-powerlink-keywords.xlsx')

wb = Workbook()

# ───────── 공통 스타일 ─────────
HEADER_FILL = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ORANGE_FILL = PatternFill(start_color='E8732A', end_color='E8732A', fill_type='solid')
ORANGE_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
SUB_FILL = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')
ZEBRA_FILL = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
BODY_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', size=10, bold=True)
THIN = Side(border_style='thin', color='E8E8E8')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

def style_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row, end_row, cols, zebra=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if not cell.alignment.horizontal:
                cell.alignment = LEFT
            if zebra and (r - start_row) % 2 == 1:
                cell.fill = ZEBRA_FILL

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═════════════════════════════════════════════════
#  시트 1: 키워드 마스터
# ═════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '1.키워드 마스터'

headers1 = ['캠페인', '광고그룹', '키워드', '추천 입찰가(원)', '우선순위', '카피ID', '랜딩 URL']
ws1.append(headers1)
style_header(ws1, 1, len(headers1))
ws1.row_dimensions[1].height = 28

# 데이터
keywords = [
    # ① 아이폰 일반
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 17 수리', 1200, '★최우선', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 17 액정 수리', 1500, '★최우선', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 17 배터리 교체', 1000, '★최우선', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 16 수리', 1000, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 16 액정 수리', 1200, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 16 배터리 교체', 900, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 15 액정 수리', 900, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 15 배터리 교체', 800, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 14 액정 수리', 700, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 14 배터리 교체', 700, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 13 액정 수리', 600, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 13 배터리 교체', 600, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 12 수리', 500, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 11 수리', 500, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 SE 수리', 500, '보조', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 충전포트 수리', 800, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 후면유리 교체', 700, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 카메라 수리', 700, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 침수 복구', 800, '중간', '카피1-A', '/'),
    ('① 아이폰 수리', '1-A 아이폰 일반', '아이폰 화면 깨짐', 700, '중간', '카피1-A', '/'),
    # ① 아이폰 가산
    ('① 아이폰 수리', '1-B 아이폰 가산', '가산 아이폰 수리', 400, '★최우선', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '가산 아이폰 액정', 400, '★최우선', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '가산 아이폰 배터리', 350, '중간', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '가산디지털단지 아이폰 수리', 450, '★최우선', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '가산디지털단지 아이폰 액정', 400, '중간', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '독산동 아이폰 수리', 300, '중간', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '독산동 아이폰 액정', 300, '중간', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '시흥동 아이폰 수리', 300, '보조', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '금천구 아이폰 수리', 300, '보조', '카피1-B', 'articles/repair-gasan-iphone.html'),
    ('① 아이폰 수리', '1-B 아이폰 가산', '가리봉동 아이폰 수리', 250, '보조', '카피1-B', 'articles/repair-gasan-iphone.html'),
    # ① 아이폰 신림
    ('① 아이폰 수리', '1-C 아이폰 신림', '신림 아이폰 수리', 400, '★최우선', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '신림 아이폰 액정', 400, '★최우선', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '신림 아이폰 배터리', 350, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '신림동 아이폰 수리', 350, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '봉천동 아이폰 수리', 300, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '봉천동 아이폰 액정', 300, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '서울대입구 아이폰 수리', 350, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '서울대입구 아이폰 액정', 350, '중간', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '신대방 아이폰 수리', 300, '보조', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '관악구 아이폰 수리', 300, '보조', '카피1-C', '/#locations'),
    ('① 아이폰 수리', '1-C 아이폰 신림', '사당 아이폰 수리', 300, '보조', '카피1-C', '/#locations'),
    # ① 아이폰 목동
    ('① 아이폰 수리', '1-D 아이폰 목동', '목동 아이폰 수리', 450, '★최우선', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '목동 아이폰 액정', 450, '★최우선', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '목동 아이폰 배터리', 400, '중간', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '목동역 아이폰 수리', 400, '★최우선', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '양천구 아이폰 수리', 400, '중간', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '양천구 아이폰 액정', 400, '중간', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '신정동 아이폰 수리', 300, '보조', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '등촌동 아이폰 수리', 300, '보조', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '화곡동 아이폰 수리', 350, '중간', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '화곡동 아이폰 액정', 350, '중간', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '신월동 아이폰 수리', 300, '보조', '카피1-D', '/#locations'),
    ('① 아이폰 수리', '1-D 아이폰 목동', '영등포 아이폰 수리', 400, '중간', '카피1-D', '/#locations'),
    # ② 아이패드 일반
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 액정 수리', 1200, '★최우선', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 배터리 교체', 1000, '★최우선', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 화면 깨짐', 900, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 충전 안됨', 800, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 프로 액정 수리', 1200, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 프로 배터리', 1000, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 미니 수리', 800, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 에어 수리', 800, '중간', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 침수', 700, '보조', '카피2-A', '/'),
    ('② 아이패드 수리', '2-A 아이패드 일반', '아이패드 카메라 수리', 600, '보조', '카피2-A', '/'),
    # ② 아이패드 가산/신림/목동
    ('② 아이패드 수리', '2-B 아이패드 가산', '가산 아이패드 수리', 400, '중간', '카피2-B', '/#locations'),
    ('② 아이패드 수리', '2-B 아이패드 가산', '가산 아이패드 액정', 400, '중간', '카피2-B', '/#locations'),
    ('② 아이패드 수리', '2-B 아이패드 가산', '독산동 아이패드 수리', 300, '보조', '카피2-B', '/#locations'),
    ('② 아이패드 수리', '2-B 아이패드 가산', '가산디지털단지 아이패드', 350, '중간', '카피2-B', '/#locations'),
    ('② 아이패드 수리', '2-C 아이패드 신림', '신림 아이패드 수리', 400, '중간', '카피2-C', '/#locations'),
    ('② 아이패드 수리', '2-C 아이패드 신림', '봉천동 아이패드 수리', 300, '보조', '카피2-C', '/#locations'),
    ('② 아이패드 수리', '2-C 아이패드 신림', '관악구 아이패드 수리', 300, '보조', '카피2-C', '/#locations'),
    ('② 아이패드 수리', '2-C 아이패드 신림', '서울대입구 아이패드', 300, '보조', '카피2-C', '/#locations'),
    ('② 아이패드 수리', '2-D 아이패드 목동', '목동 아이패드 수리', 450, '★최우선', '카피2-D', '/#locations'),
    ('② 아이패드 수리', '2-D 아이패드 목동', '양천구 아이패드 수리', 400, '중간', '카피2-D', '/#locations'),
    ('② 아이패드 수리', '2-D 아이패드 목동', '화곡동 아이패드 수리', 350, '보조', '카피2-D', '/#locations'),
    ('② 아이패드 수리', '2-D 아이패드 목동', '목동역 아이패드 수리', 400, '중간', '카피2-D', '/#locations'),
    # ③ 애플워치 일반
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 수리', 1200, '★최우선', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 배터리 교체', 1000, '★최우선', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 액정 교체', 1000, '★최우선', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 후면유리', 800, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 침수', 700, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', 'Apple Watch 수리', 1000, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 시리즈 10 수리', 1200, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 시리즈 9 수리', 900, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 시리즈 8 수리', 700, '보조', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 Ultra 수리', 900, '중간', '카피3-A', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-A 애플워치 일반', '애플워치 SE 수리', 700, '보조', '카피3-A', 'articles/hub-watch.html'),
    # ③ 애플워치 가산/신림/목동
    ('③ 애플워치 수리', '3-B 애플워치 가산', '가산 애플워치 수리', 400, '중간', '카피3-B', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-B 애플워치 가산', '가산 애플워치 배터리', 350, '중간', '카피3-B', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-B 애플워치 가산', '가산디지털단지 애플워치', 350, '보조', '카피3-B', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-C 애플워치 신림', '신림 애플워치 수리', 400, '중간', '카피3-C', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-C 애플워치 신림', '봉천동 애플워치 수리', 300, '보조', '카피3-C', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-C 애플워치 신림', '관악구 애플워치 수리', 300, '보조', '카피3-C', 'articles/hub-watch.html'),
    ('③ 애플워치 수리', '3-D 애플워치 목동', '목동 애플워치 수리', 450, '★최우선', '카피3-D', 'articles/repair-mokdong-watch.html'),
    ('③ 애플워치 수리', '3-D 애플워치 목동', '목동 애플워치 배터리', 400, '중간', '카피3-D', 'articles/repair-mokdong-watch.html'),
    ('③ 애플워치 수리', '3-D 애플워치 목동', '양천구 애플워치 수리', 400, '중간', '카피3-D', 'articles/repair-mokdong-watch.html'),
    # ④ 애플펜슬 일반
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 수리', 800, '★최우선', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 2세대 수리', 900, '★최우선', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 배터리 교체', 700, '★최우선', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 충전 안됨', 600, '★최우선', '카피4-A', 'articles/applepencil-gen2-repair-diagnosis.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 인식 안됨', 500, '중간', '카피4-A', 'articles/applepencil-gen2-repair-diagnosis.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 리퍼', 600, '중간', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 교체 가격', 700, '중간', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', 'Apple Pencil 수리', 700, '중간', '카피4-A', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 고장', 500, '중간', '카피4-A', 'articles/applepencil-gen2-repair-diagnosis.html'),
    ('④ 애플펜슬 수리', '4-A 애플펜슬 일반', '애플펜슬 필기 안됨', 450, '보조', '카피4-A', 'articles/applepencil-gen2-repair-diagnosis.html'),
    # ④ 애플펜슬 가산/신림/목동
    ('④ 애플펜슬 수리', '4-B 애플펜슬 가산', '가산 애플펜슬 수리', 400, '중간', '카피4-B', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-B 애플펜슬 가산', '가산디지털단지 애플펜슬', 400, '중간', '카피4-B', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-C 애플펜슬 신림', '신림 애플펜슬 수리', 400, '중간', '카피4-C', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-C 애플펜슬 신림', '봉천동 애플펜슬', 300, '보조', '카피4-C', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-D 애플펜슬 목동', '목동 애플펜슬 수리', 450, '★최우선', '카피4-D', 'articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬 수리', '4-D 애플펜슬 목동', '양천구 애플펜슬', 400, '중간', '카피4-D', 'articles/applepencil-gen2-repair-service.html'),
]

for kw in keywords:
    ws1.append(kw)

style_body(ws1, 2, ws1.max_row, len(headers1))

# 우선순위 칼럼 색상
for r in range(2, ws1.max_row + 1):
    pri = ws1.cell(row=r, column=5).value
    cell = ws1.cell(row=r, column=5)
    cell.alignment = CENTER
    if pri == '★최우선':
        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color='DC2626')
    elif pri == '중간':
        cell.fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    elif pri == '보조':
        cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    # 입찰가 우측 정렬
    ws1.cell(row=r, column=4).alignment = RIGHT

set_widths(ws1, [16, 18, 25, 14, 10, 10, 50])
ws1.freeze_panes = 'A2'

# ═════════════════════════════════════════════════
#  시트 2: 카피 세트
# ═════════════════════════════════════════════════
ws2 = wb.create_sheet('2.카피 세트')

headers2 = ['카피ID', '광고그룹', '제목 (15자)', '설명1 (45자)', '설명2 (45자)', '랜딩 URL']
ws2.append(headers2)
style_header(ws2, 1, len(headers2))
ws2.row_dimensions[1].height = 28

copies = [
    ('카피1-A', '아이폰 일반', '아이폰 수리 다올리페어', '당일 30분 완료ㆍ수리 실패 시 비용 0원', '가산ㆍ신림ㆍ목동 3지점ㆍ3개월 무상 A/S', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('카피1-B', '아이폰 가산', '가산 아이폰 수리 즉시', '가산디지털단지 5분 거리ㆍ당일 30분 완료', '점심 맡기고 퇴근에 수령ㆍ수리 실패시 0원', 'https://xn--2j1bq2k97kxnah86c.com/articles/repair-gasan-iphone.html'),
    ('카피1-C', '아이폰 신림', '신림 아이폰 수리 당일', '신림역 도보 거리ㆍ액정·배터리 당일 교체', '관악구 거주자 다수 후기ㆍ수리 실패시 0원', 'https://xn--2j1bq2k97kxnah86c.com/#locations'),
    ('카피1-D', '아이폰 목동', '목동 아이폰 수리 당일', '목동역 도보 5분ㆍ당일 30분 완료', '양천구 후기 다수ㆍ3개월 A/Sㆍ수리 실패시 0원', 'https://xn--2j1bq2k97kxnah86c.com/#locations'),
    ('카피2-A', '아이패드 일반', '아이패드 액정 당일 수리', '액정·배터리·충전포트 30분 내 완료', '데이터 보존 100%ㆍ수리 실패시 0원ㆍ서울 3지점', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('카피2-B', '아이패드 가산', '가산 아이패드 수리', '가산디지털단지 5분ㆍ액정·배터리 당일 교체', '데이터 보존ㆍ수리 실패시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/#locations'),
    ('카피2-C', '아이패드 신림', '신림 아이패드 수리', '신림역 도보 거리ㆍ액정·배터리 당일 교체', '데이터 보존ㆍ수리 실패시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/#locations'),
    ('카피2-D', '아이패드 목동', '목동 아이패드 수리', '목동역 도보 5분ㆍ액정·배터리 당일 교체', '양천구 거주자 다수ㆍ수리 실패시 0원', 'https://xn--2j1bq2k97kxnah86c.com/#locations'),
    ('카피3-A', '애플워치 일반', '애플워치 배터리·액정', 'Series 4부터 Ultra까지 모두 수리 가능', '당일 완료ㆍ서울 3지점ㆍ3개월 무상 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/hub-watch.html'),
    ('카피3-B', '애플워치 가산', '가산 애플워치 수리', '가산디지털단지 5분ㆍ배터리·액정 당일', 'Series 4~Ultraㆍ수리 실패시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/hub-watch.html'),
    ('카피3-C', '애플워치 신림', '신림 애플워치 수리', '신림 도보 거리ㆍ배터리·액정 당일 교체', 'Series 4~Ultraㆍ수리 실패시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/hub-watch.html'),
    ('카피3-D', '애플워치 목동', '목동 애플워치 수리', '목동역 도보 5분ㆍ배터리·액정 당일 교체', '양천구 거주자 다수 후기ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/repair-mokdong-watch.html'),
    ('카피4-A', '애플펜슬 일반', '애플펜슬 2세대 8만원 당일', '공식 AS 16.9만원→다올리페어 8만원 당일', '1:1 리퍼 교체ㆍ수리 불가시 0원ㆍ서울 3지점', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
    ('카피4-B', '애플펜슬 가산', '가산 애플펜슬 8만원', '공식 16.9만→8만원ㆍ가산에서 당일 교체', '1:1 리퍼ㆍ수리 불가시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
    ('카피4-C', '애플펜슬 신림', '신림 애플펜슬 8만원', '공식 16.9만→8만원ㆍ신림에서 당일 교체', '1:1 리퍼ㆍ수리 불가시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
    ('카피4-D', '애플펜슬 목동', '목동 애플펜슬 8만원', '공식 16.9만→8만원ㆍ목동에서 당일 교체', '1:1 리퍼ㆍ수리 불가시 0원ㆍ3개월 A/S', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
]

for cp in copies:
    ws2.append(cp)

style_body(ws2, 2, ws2.max_row, len(headers2))

# 카피ID 칼럼은 가운데정렬 + bold
for r in range(2, ws2.max_row + 1):
    cell = ws2.cell(row=r, column=1)
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.fill = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')
    # 행 높이
    ws2.row_dimensions[r].height = 32

set_widths(ws2, [12, 18, 24, 38, 38, 60])
ws2.freeze_panes = 'A2'

# ═════════════════════════════════════════════════
#  시트 3: 일별 예산 분배
# ═════════════════════════════════════════════════
ws3 = wb.create_sheet('3.예산 분배')

headers3 = ['캠페인', '광고그룹', '일 예산(원)', '월 환산(원)', '비고']
ws3.append(headers3)
style_header(ws3, 1, len(headers3))
ws3.row_dimensions[1].height = 28

budget = [
    ('④ 애플펜슬 ★★★', '4-A 애플펜슬 일반', 9000, 270000, '차별화 무기 — 공식대비 9만원 절감'),
    ('④ 애플펜슬 ★★★', '4-B 애플펜슬 가산', 2000, 60000, ''),
    ('④ 애플펜슬 ★★★', '4-C 애플펜슬 신림', 2000, 60000, ''),
    ('④ 애플펜슬 ★★★', '4-D 애플펜슬 목동', 2000, 60000, ''),
    ('① 아이폰 ★★★', '1-A 아이폰 일반(모델별)', 9000, 270000, '메인 매출 — 모델별 키워드 위주'),
    ('① 아이폰 ★★★', '1-B 아이폰 가산', 2000, 60000, ''),
    ('① 아이폰 ★★★', '1-C 아이폰 신림', 2000, 60000, ''),
    ('① 아이폰 ★★★', '1-D 아이폰 목동', 2000, 60000, ''),
    ('② 아이패드 ★★', '2-A 아이패드 일반', 4000, 120000, ''),
    ('② 아이패드 ★★', '2-B 아이패드 가산', 2000, 60000, ''),
    ('② 아이패드 ★★', '2-C 아이패드 신림', 2000, 60000, ''),
    ('② 아이패드 ★★', '2-D 아이패드 목동', 2000, 60000, ''),
    ('③ 애플워치 ★★', '3-A 애플워치 일반', 4000, 120000, ''),
    ('③ 애플워치 ★★', '3-B 애플워치 가산', 2000, 60000, ''),
    ('③ 애플워치 ★★', '3-C 애플워치 신림', 2000, 60000, ''),
    ('③ 애플워치 ★★', '3-D 애플워치 목동', 2000, 60000, ''),
]

for b in budget:
    ws3.append(b)

# 합계 행
total_day = sum(b[2] for b in budget)
total_month = sum(b[3] for b in budget)
ws3.append(['합계', '16개 광고그룹', total_day, total_month, '하루 5만원 / 월 150만원'])

style_body(ws3, 2, ws3.max_row, len(headers3))

# 합계 행 강조
last = ws3.max_row
for c in range(1, len(headers3) + 1):
    cell = ws3.cell(row=last, column=c)
    cell.fill = ORANGE_FILL
    cell.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')

# 숫자 칼럼 우측 정렬
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=3).alignment = RIGHT
    ws3.cell(row=r, column=4).alignment = RIGHT
    ws3.cell(row=r, column=3).number_format = '#,##0'
    ws3.cell(row=r, column=4).number_format = '#,##0'

set_widths(ws3, [16, 22, 14, 14, 40])
ws3.freeze_panes = 'A2'

# ═════════════════════════════════════════════════
#  시트 4: 1주차 점검 체크리스트
# ═════════════════════════════════════════════════
ws4 = wb.create_sheet('4.점검 체크리스트')

headers4 = ['요일', '시간', '체크 항목', '확인']
ws4.append(headers4)
style_header(ws4, 1, len(headers4))
ws4.row_dimensions[1].height = 28

checklist = [
    ('월요일', '오전', '모든 캠페인 ON 확인', ''),
    ('월요일', '오전', '확장 소재 4종 모두 등록 확인 (서브링크/이미지/전화/추가설명)', ''),
    ('월요일', '오전', '네이버 플레이스 3지점 연동 확인', ''),
    ('월요일', '저녁', '광고 노출 시작 확인 (5분 이상 노출 0건이면 입찰가 ↑)', ''),
    ('수요일', '저녁', '광고그룹별 노출수 확인', ''),
    ('수요일', '저녁', 'CTR 5% 미만 키워드 표시 → 카피 점검', ''),
    ('수요일', '저녁', '노출 100회 미만 키워드 입찰가 +20%', ''),
    ('금요일', '저녁', '키워드별 클릭당 단가(CPC) 확인', ''),
    ('금요일', '저녁', 'CPC 너무 비싼 키워드(예상 대비 +50%) 체크', ''),
    ('금요일', '저녁', '실제 견적 문의 발생 키워드 정리 (시트에 기록)', ''),
    ('금요일', '저녁', '문의 0건 + 클릭 많은 키워드 일시 중지 검토', ''),
    ('일요일', '저녁', '1주 데이터 종합 정리', ''),
    ('일요일', '저녁', '잘 되는 광고그룹 예산 ↑ (2~3배)', ''),
    ('일요일', '저녁', '안 되는 광고그룹 카피 수정 또는 정지', ''),
    ('일요일', '저녁', '2주차 새 키워드 추가 후보 작성', ''),
]

for ck in checklist:
    ws4.append(ck)

style_body(ws4, 2, ws4.max_row, len(headers4))

# 요일 칼럼 강조
for r in range(2, ws4.max_row + 1):
    cell = ws4.cell(row=r, column=1)
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.fill = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')
    ws4.cell(row=r, column=2).alignment = CENTER
    ws4.cell(row=r, column=4).alignment = CENTER

set_widths(ws4, [12, 10, 60, 8])
ws4.freeze_panes = 'A2'

# ═════════════════════════════════════════════════
#  시트 5: 확장 소재 일괄 등록
# ═════════════════════════════════════════════════
ws5 = wb.create_sheet('5.확장 소재')

headers5 = ['소재 종류', '내용', '비고']
ws5.append(headers5)
style_header(ws5, 1, len(headers5))
ws5.row_dimensions[1].height = 28

extras = [
    ('서브링크 1', '아이폰 액정 수리 → https://xn--2j1bq2k97kxnah86c.com', '메인 페이지'),
    ('서브링크 2', '배터리 교체 → https://xn--2j1bq2k97kxnah86c.com', '메인 페이지'),
    ('서브링크 3', '애플워치 수리 → /articles/hub-watch.html', '애플워치 허브'),
    ('서브링크 4', '애플펜슬 8만원 → /articles/applepencil-gen2-repair-service.html', '펜슬 차별화 페이지'),
    ('이미지 (썸네일)', '매장 외관 또는 수리 작업 장면 1장 (텍스트 없는 깔끔한 사진)', '권장 사이즈 1200x900'),
    ('전화번호', '010-9904-1535', '모바일 클릭 → 통화 즉시 연결'),
    ('추가 설명문구', '대한민국 1호 디바이스 예방 마스터ㆍ당일 수리 전문점', '15자 내외'),
    ('네이버 플레이스 1', '가산점 — 영업시간·전화·사진·후기 모두 채워둘 것', '광고 시작 전 점검 필수'),
    ('네이버 플레이스 2', '신림점 — 영업시간·전화·사진·후기 모두 채워둘 것', '광고 시작 전 점검 필수'),
    ('네이버 플레이스 3', '목동점 — 영업시간·전화·사진·후기 모두 채워둘 것', '광고 시작 전 점검 필수'),
]

for e in extras:
    ws5.append(e)

style_body(ws5, 2, ws5.max_row, len(headers5))

# 소재 종류 강조
for r in range(2, ws5.max_row + 1):
    cell = ws5.cell(row=r, column=1)
    cell.font = BOLD_FONT
    cell.fill = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')

set_widths(ws5, [22, 60, 28])
ws5.freeze_panes = 'A2'

# ─── 저장 ───
wb.save(OUT)
print('[OK] saved:', OUT)
print('   sheets: 5 / keywords:', len(keywords), '/ copies:', len(copies))
