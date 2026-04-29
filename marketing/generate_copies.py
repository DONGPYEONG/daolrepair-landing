# -*- coding: utf-8 -*-
"""
다올리페어 네이버 파워링크 — 광고 카피 + 확장소재 정리 엑셀
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'naver-powerlink-copies.xlsx')

# 스타일
HEADER_FILL = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ORANGE_FILL = PatternFill(start_color='E8732A', end_color='E8732A', fill_type='solid')
ORANGE_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
SUB_FILL = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')
ZEBRA = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
HIGHLIGHT_FILL = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')

BODY = Font(name='맑은 고딕', size=10)
BOLD = Font(name='맑은 고딕', size=10, bold=True)
TITLE_FONT = Font(name='맑은 고딕', size=12, bold=True, color='E8732A')

THIN = Side(border_style='thin', color='E8E8E8')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, n, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def style_body(ws, start, end, n, highlight_first=False):
    for r in range(start, end + 1):
        for c in range(1, n + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = LEFT if c >= 2 else CENTER
            if (r - start) % 2 == 1:
                cell.fill = ZEBRA
        if highlight_first:
            cell = ws.cell(row=r, column=1)
            cell.font = BOLD
            cell.fill = SUB_FILL


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════
# 시트 1: 카테고리별 광고 카피
# ═════════════════════════════════════════════════
ws1 = wb.create_sheet('1.카테고리별 광고')

# 안내 행
ws1.append(['카테고리', '버전', '제목 (15자)', '설명1 (45자)', '설명2 (45자)', '표시 URL', '연결 URL'])
style_header(ws1, 1, 7)

copies = [
    # 아이폰
    ('① 아이폰 ★★★', '일반', '아이폰 수리 다올리페어',
     '당일 30분 완료ㆍ수리 실패 시 비용 0원 보장',
     '가산ㆍ신림ㆍ목동 3개 직영점ㆍ3개월 무상 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('① 아이폰', '액정 강조', '아이폰 액정 30분 수리',
     '당일 30분ㆍ정품/사제 액정 모두 가능',
     '데이터 100% 보존ㆍ수리 실패시 0원ㆍ3개월 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('① 아이폰', '아이폰17', '아이폰17 액정 당일 수리',
     '아이폰17 전 모델 30분 액정 교체ㆍ당일 가능',
     '데이터 손실 0%ㆍ수리 실패시 비용 0원',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('① 아이폰', '배터리', '아이폰 배터리 당일 교체',
     '정품 배터리ㆍ80% 미만 즉시 교체ㆍ30분 완료',
     '수리 실패시 비용 0원ㆍ서울 3지점ㆍ3개월 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    # 아이패드
    ('② 아이패드 ★★', '일반', '아이패드 액정 당일 수리',
     '액정ㆍ배터리ㆍ충전포트 30분 내 완료',
     '데이터 100% 보존ㆍ수리 실패시 0원ㆍ서울 3지점',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('② 아이패드', '프로 강조', '아이패드프로 수리 전문',
     'M4ㆍM2 모델 당일 액정 교체 가능',
     '정품 부품ㆍ수리 실패시 0원ㆍ3개월 무상 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('② 아이패드', '액정', '아이패드 액정 깨짐',
     '당일 수리 가능ㆍ프로/에어/미니 모든 모델',
     '정품 부품ㆍ수리 실패시 0원ㆍ3개월 무상 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    # 애플워치
    ('③ 애플워치 ★★', '일반', '애플워치 배터리·액정',
     'Series 4부터 Ultra까지 모두 수리 가능',
     '당일 완료ㆍ서울 3지점ㆍ3개월 무상 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('③ 애플워치', '액정', '애플워치 액정 교체',
     '모든 시리즈 액정 당일 교체ㆍ정품 디스플레이',
     '수리 실패 시 0원ㆍ가산/신림/목동 3개 직영점',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    ('③ 애플워치', '배터리', '애플워치 배터리 교체',
     'Series 4~Ultraㆍ당일 배터리 교체 가능',
     '정품 배터리ㆍ수리 실패시 0원ㆍ3개월 A/S',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com'),
    # 애플펜슬
    ('④ 애플펜슬 ★★★', '메인 (가격 강조)', '애플펜슬2세대 8만원',
     '공식 AS 16.9만원→다올리페어 8만원 당일',
     '1:1 리퍼 교체ㆍ수리 불가시 0원ㆍ서울 3지점',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬', '증상별', '애플펜슬 수리 8만원',
     '충전 안됨ㆍ인식 불량ㆍ당일 8만원 해결',
     '1:1 리퍼ㆍ공식 대비 약 9만원 절감ㆍ당일',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
    ('④ 애플펜슬', '배터리', '애플펜슬 배터리 교체',
     '2세대 1:1 리퍼ㆍ8만원 당일 교체',
     '공식 AS 16.9만원 대비 9만원 절감ㆍ서울 3지점',
     '다올리페어.com', 'https://xn--2j1bq2k97kxnah86c.com/articles/applepencil-gen2-repair-service.html'),
]

for row in copies:
    ws1.append(row)

style_body(ws1, 2, ws1.max_row, 7, highlight_first=True)
set_widths(ws1, [16, 16, 24, 38, 38, 18, 50])
ws1.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
# 시트 2: 지역별 광고 카피
# ═════════════════════════════════════════════════
ws2 = wb.create_sheet('2.지역별 광고')
ws2.append(['지점', '버전', '제목 (15자)', '설명1 (45자)', '설명2 (45자)', '연결 URL'])
style_header(ws2, 1, 6)

region_copies = [
    # 가산
    ('가산점', '메인', '가산 아이폰 수리 즉시',
     '가산디지털단지 5분 거리ㆍ당일 30분 완료',
     '점심 맡기고 퇴근에 수령ㆍ수리 실패시 0원',
     'https://xn--2j1bq2k97kxnah86c.com/articles/repair-gasan-iphone.html'),
    ('가산점', '액정 강조', '가산 액정 당일 교체',
     '직장인 점심시간 수리ㆍ퇴근 전 수령 가능',
     '데이터 보존 100%ㆍ3개월 A/Sㆍ금천구 직영',
     'https://xn--2j1bq2k97kxnah86c.com/articles/repair-gasan-iphone.html'),
    ('가산점', '독산동', '독산동 아이폰 수리',
     '독산역 5분ㆍ당일 30분 완료ㆍ직장인 환영',
     '수리 실패시 0원ㆍ3개월 A/Sㆍ퇴근길 수령',
     'https://xn--2j1bq2k97kxnah86c.com/articles/repair-gasan-iphone.html'),
    # 신림
    ('신림점', '메인', '신림 아이폰 수리 당일',
     '신림역 도보 5분ㆍ액정·배터리 당일 교체',
     '관악구 거주자 다수 후기ㆍ수리 실패시 0원',
     'https://xn--2j1bq2k97kxnah86c.com'),
    ('신림점', '봉천동', '봉천동 아이폰 액정',
     '신림 직영점ㆍ아이폰17~11 모든 모델 가능',
     '정품 부품ㆍ당일 30분ㆍ3개월 무상 A/S',
     'https://xn--2j1bq2k97kxnah86c.com'),
    ('신림점', '서울대입구', '서울대입구 아이폰 수리',
     '신림역 직영점 5분 거리ㆍ학생/직장인 환영',
     '당일 완료ㆍ수리 실패시 0원ㆍ3개월 A/S',
     'https://xn--2j1bq2k97kxnah86c.com'),
    # 목동
    ('목동점', '메인', '목동 아이폰 수리 당일',
     '목동역 도보 5분ㆍ당일 30분 완료',
     '양천구 후기 다수ㆍ3개월 A/Sㆍ수리 실패시 0원',
     'https://xn--2j1bq2k97kxnah86c.com'),
    ('목동점', '화곡동', '화곡동 액정 수리',
     '양천구 직영점ㆍ아이폰·아이패드·워치 가능',
     '정품 부품ㆍ당일 완료ㆍ수리 실패시 0원',
     'https://xn--2j1bq2k97kxnah86c.com'),
    ('목동점', '신정동', '신정동 아이폰 수리',
     '목동 직영점 도보 거리ㆍ당일 액정·배터리',
     '데이터 보존ㆍ수리 실패시 0원ㆍ3개월 A/S',
     'https://xn--2j1bq2k97kxnah86c.com'),
]

for row in region_copies:
    ws2.append(row)

style_body(ws2, 2, ws2.max_row, 6, highlight_first=True)
set_widths(ws2, [12, 14, 24, 38, 38, 50])
ws2.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
# 시트 3: 확장소재 (서브링크/추가설명/이미지/플레이스)
# ═════════════════════════════════════════════════
ws3 = wb.create_sheet('3.확장소재')
ws3.append(['소재 종류', '항목', '내용', '비고'])
style_header(ws3, 1, 4)

extras = [
    # 서브링크 4개
    ('서브링크', '1번', '아이폰 액정 수리 → 30분 당일 완료ㆍ수리 실패시 0원',
     '연결: 메인 페이지'),
    ('서브링크', '2번', '배터리 교체 → 정품 배터리ㆍ당일ㆍ3개월 A/S',
     '연결: 메인 페이지'),
    ('서브링크', '3번', '애플워치 수리 → Series 4~Ultra 모두 가능',
     '연결: hub-watch.html'),
    ('서브링크', '4번', '애플펜슬 8만원 → 공식 16.9만→8만원 당일 교체',
     '연결: applepencil-gen2-repair-service.html'),
    # 추가설명
    ('추가설명', '1번', '대한민국 1호 디바이스 예방 마스터',
     '회전 노출 (3개까지 등록 가능)'),
    ('추가설명', '2번', '가산ㆍ신림ㆍ목동 직영점 3개 운영',
     '회전 노출'),
    ('추가설명', '3번', '수리 실패 시 비용 0원ㆍ3개월 무상 A/S',
     '회전 노출'),
    # 전화번호
    ('전화번호', '대표번호', '010-9904-1535',
     '모바일 클릭 → 통화 연결'),
    ('전화번호', '표시 텍스트', '지금 통화로 견적 받기',
     '클릭 유도 문구'),
    # 이미지
    ('이미지', '권장 사이즈', '1200×900 (4:3 비율)',
     '이미지 확장소재용 썸네일'),
    ('이미지', '추천 종류', '매장 외관 / 수리 작업 장면 / 수리 전후 비교',
     '텍스트 없는 깔끔한 사진 권장'),
    ('이미지', '금지 사항', '텍스트 과다, 어두운 배경, 사람 얼굴 노출',
     '카카오/네이버 가이드 위반 시 노출 거부'),
    # 네이버 플레이스
    ('네이버 플레이스', '가산점', '영업시간·전화·사진 5장+·후기 답글 등록',
     '필수 ★ 등록 안 하면 광고 효과 30% 감소'),
    ('네이버 플레이스', '신림점', '영업시간·전화·사진 5장+·후기 답글 등록',
     '필수 ★'),
    ('네이버 플레이스', '목동점', '영업시간·전화·사진 5장+·후기 답글 등록',
     '필수 ★ 위치 기반 노출'),
    # 추가 추천 확장소재
    ('블로그 리뷰', '연동', '다올리페어 후기 블로그 글 등록',
     '검색 시 블로그 후기 같이 노출'),
    ('네이버 톡톡', '연동', '톡톡 채팅 즉시 응대',
     '빠른 상담 유입 (선택)'),
    ('가격 표시', '예시', '애플펜슬 2세대 80,000원',
     '명확한 가격으로 신뢰도 ↑'),
]

for row in extras:
    ws3.append(row)

style_body(ws3, 2, ws3.max_row, 4, highlight_first=True)
set_widths(ws3, [16, 12, 50, 32])
ws3.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
# 시트 4: 후킹 표현 모음 (자유 활용)
# ═════════════════════════════════════════════════
ws4 = wb.create_sheet('4.후킹 표현 모음')
ws4.append(['카테고리', '표현', '글자수', '용도'])
style_header(ws4, 1, 4)

phrases = [
    # 시간/속도
    ('🕐 시간/속도', '당일 30분 완료', 8, '제목·설명1'),
    ('🕐 시간/속도', '점심 맡기고 퇴근에 수령', 14, '설명2'),
    ('🕐 시간/속도', '출근 전 맡기고 점심에 수령', 15, '설명2'),
    ('🕐 시간/속도', '30분이면 끝', 7, '제목'),
    ('🕐 시간/속도', '즉시 진단ㆍ즉시 수리', 11, '설명'),
    ('🕐 시간/속도', '대기 없이 바로 접수', 10, '설명'),
    ('🕐 시간/속도', '오늘 깨졌어도 오늘 받아가세요', 16, '설명2'),
    ('🕐 시간/속도', '당일 예약 가능ㆍ즉시 응대', 14, '설명'),
    # 비용/가격
    ('💰 비용/가격', '공식 AS 대비 50% 절감', 13, '설명1'),
    ('💰 비용/가격', '애플펜슬 2세대 8만원 (공식 16.9만)', 19, '설명2'),
    ('💰 비용/가격', '견적 무료ㆍ수리 실패시 0원', 14, '설명'),
    ('💰 비용/가격', '숨은 비용 0원ㆍ투명 견적', 13, '설명'),
    ('💰 비용/가격', '3개월 A/S 포함 가격', 11, '설명'),
    # 신뢰/안심
    ('🛡 신뢰/안심', '대한민국 1호 디바이스 예방 마스터', 18, '설명/추가설명'),
    ('🛡 신뢰/안심', '직접 수리ㆍ외부 대행 없음', 13, '설명'),
    ('🛡 신뢰/안심', '정품 부품 사용', 8, '설명'),
    ('🛡 신뢰/안심', '데이터 100% 보존', 10, '설명'),
    ('🛡 신뢰/안심', '수리 실패 시 비용 0원', 12, '설명1·2'),
    ('🛡 신뢰/안심', '3개월 무상 A/S 보증', 11, '설명2'),
    ('🛡 신뢰/안심', '서울 3개 직영점 운영', 11, '설명'),
    # 긴급/공감
    ('🚨 긴급/공감', '오늘 깨졌어도 오늘 받아가세요', 16, '설명2'),
    ('🚨 긴급/공감', '출근 전ㆍ퇴근 후 방문 가능', 14, '설명'),
    ('🚨 긴급/공감', '당일 예약 가능ㆍ즉시 응대', 14, '설명'),
    ('🚨 긴급/공감', '배터리 80% 미만 즉시 교체', 14, '설명'),
    ('🚨 긴급/공감', '화면 줄/얼룩 즉시 점검', 12, '설명'),
]

for row in phrases:
    ws4.append(row)

style_body(ws4, 2, ws4.max_row, 4, highlight_first=True)
for r in range(2, ws4.max_row + 1):
    ws4.cell(row=r, column=3).alignment = CENTER
    ws4.cell(row=r, column=4).alignment = CENTER

set_widths(ws4, [16, 38, 10, 18])
ws4.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
# 시트 5: 등록 전 체크리스트 + 운영 팁
# ═════════════════════════════════════════════════
ws5 = wb.create_sheet('5.체크리스트+운영팁')
ws5.append(['단계', '항목', '확인'])
style_header(ws5, 1, 3)

checklist = [
    # 등록 전
    ('등록 전', '네이버 플레이스 가산점 등록 + 사진·후기·영업시간', ''),
    ('등록 전', '네이버 플레이스 신림점 등록 + 사진·후기·영업시간', ''),
    ('등록 전', '네이버 플레이스 목동점 등록 + 사진·후기·영업시간', ''),
    ('등록 전', '매장 사진 1장 이상 준비 (확장소재 이미지용)', ''),
    ('등록 전', '전화번호 010-9904-1535 표기 통일', ''),
    ('등록 전', '랜딩 페이지 다올리페어.com 정상 작동 확인', ''),
    ('등록 전', '카톡 채널 (https://pf.kakao.com/_xfRNMX) 친구 추가 가능 확인', ''),
    # 광고 등록
    ('광고 등록', '캠페인 4개 생성 (아이폰/아이패드/애플워치/애플펜슬)', ''),
    ('광고 등록', '각 캠페인에 광고그룹 4개씩 (일반/가산/신림/목동)', ''),
    ('광고 등록', '카테고리별 카피 세트 등록 (시트 1 참조)', ''),
    ('광고 등록', '지역별 카피 세트 등록 (시트 2 참조)', ''),
    ('광고 등록', '확장소재 4종 등록 (시트 3 참조)', ''),
    ('광고 등록', '일 예산 5만원 설정', ''),
    ('광고 등록', '모바일 입찰 가중치 +30% 설정', ''),
    # 1주차 운영
    ('1주차 운영', '월요일: 광고 ON + 노출 확인', ''),
    ('1주차 운영', '수요일 저녁: 중간 점검 (입찰가 조정)', ''),
    ('1주차 운영', '금요일 저녁: 키워드별 성과 확인', ''),
    ('1주차 운영', '일요일 저녁: 1주 데이터 정리 + 2주차 계획', ''),
    # 2주차 이후
    ('2주차+', 'CTR 5% 미만 키워드 카피 교체', ''),
    ('2주차+', '잘 되는 키워드 입찰가 +20% 인상', ''),
    ('2주차+', 'A/B 테스트 (카피 2개 → 1주일 후 잘 되는 거 남기기)', ''),
    ('2주차+', '시간대별 입찰 가중치 조정 (지점별 피크 시간)', ''),
]

for row in checklist:
    ws5.append(row)

style_body(ws5, 2, ws5.max_row, 3, highlight_first=True)
for r in range(2, ws5.max_row + 1):
    ws5.cell(row=r, column=3).alignment = CENTER
    cell = ws5.cell(row=r, column=1)
    if '등록 전' in cell.value:
        cell.fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    elif '광고 등록' in cell.value:
        cell.fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    elif '1주차' in cell.value:
        cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    elif '2주차' in cell.value:
        cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

set_widths(ws5, [12, 60, 8])
ws5.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
# 시트 0: 전체 요약
# ═════════════════════════════════════════════════
ws0 = wb.create_sheet('0.시작하기', 0)
ws0.append(['시트', '내용', '설명'])
style_header(ws0, 1, 3)

intro = [
    ('1.카테고리별 광고', '13개 광고 카피', '아이폰/아이패드/애플워치/애플펜슬 — 일반·모델별·증상별'),
    ('2.지역별 광고', '9개 지역 카피', '가산점/신림점/목동점 + 인접 지역명 강조'),
    ('3.확장소재', '서브링크·이미지·전화·플레이스 등', '한 번 세팅하면 모든 광고에 자동 적용'),
    ('4.후킹 표현 모음', '25개 자유 활용 문구', '시간/비용/신뢰/긴급 4가지 카테고리'),
    ('5.체크리스트+운영팁', '등록 전 + 운영 점검표', '단계별 빠짐없이 체크'),
]

for row in intro:
    ws0.append(row)

style_body(ws0, 2, ws0.max_row, 3, highlight_first=True)
set_widths(ws0, [22, 28, 60])

# 푸터 안내
ws0.cell(row=ws0.max_row + 2, column=1).value = '💡 사용 안내'
ws0.cell(row=ws0.max_row, column=1).font = Font(name='맑은 고딕', size=12, bold=True, color='E8732A')

guide_rows = [
    '1. 시트 1·2의 카피를 네이버 광고 시스템에 그대로 복사해서 등록',
    '2. 시트 3의 확장소재를 한 번에 세팅 (모든 광고에 자동 적용)',
    '3. 시트 5의 체크리스트로 등록 전·후 점검',
    '4. 카피 변형 필요 시 시트 4의 표현을 자유롭게 조합',
    '5. 1주일 운영 후 잘 되는 키워드만 남기고 입찰가 조정',
]

for g in guide_rows:
    ws0.cell(row=ws0.max_row + 1, column=1).value = g
    ws0.cell(row=ws0.max_row, column=1).font = Font(name='맑은 고딕', size=10)

ws0.freeze_panes = 'A2'

wb.save(OUT)
print('[OK] saved:', OUT)
