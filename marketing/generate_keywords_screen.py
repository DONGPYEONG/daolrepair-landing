# -*- coding: utf-8 -*-
"""
다올리페어 네이버 파워링크 — 액정 수리 중심 키워드 재편
- 각 광고그룹당 100개, 이 중 70%는 액정 관련
- 액정 키워드는 입찰가 +30% 할증 (2~3위 노출 목표)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'naver-powerlink-keywords-screen.xlsx')

# ═════════════════════════════════════════════════
#  키워드 풀 — 액정 중심
# ═════════════════════════════════════════════════

# 아이폰 모델 (우선순위별)
IPHONE_MODELS_HOT = ['17 Pro Max', '17 Pro', '17', '16 Pro Max', '16 Pro', '16']
IPHONE_MODELS_MID = ['15 Pro Max', '15 Pro', '15', '14 Pro Max', '14 Pro', '14', '13 Pro', '13', 'SE 3']
IPHONE_MODELS_LOW = ['12 Pro', '12', '11 Pro Max', '11 Pro', '11', 'SE 2', 'XS', 'XR', 'X']

# 액정 관련 표현 (수익 핵심)
SCREEN_TERMS_HOT = ['액정 수리', '액정 교체', '액정 깨짐', '화면 깨짐', '디스플레이 교체']
SCREEN_TERMS_MID = ['액정', '화면 수리', '화면 교체', 'OLED 교체', 'LCD 교체', '전면유리 교체',
                    '액정 터치 안됨', '액정 줄', '화면 초록줄', '화면 분홍줄', '화면 검정',
                    '액정 갈라짐', '유리 깨짐', '액정 수리비', '액정 가격']
SCREEN_TERMS_LOW = ['액정 당일', '액정 저렴', '액정 전문점', '액정 잘하는곳', '액정 정품',
                    '액정 사제', '액정 무상', '액정 긴급', '액정 출장', '액정 방문',
                    '액정 후기', '액정 보증', '액정 할인', '액정 견적']

# 기타 핵심 수리 (30%)
IPHONE_OTHER_TYPES = ['배터리 교체', '충전포트 수리', '카메라 수리', '후면유리 교체',
                      '침수 복구', '버튼 수리', '스피커 수리', '홈버튼 수리']

# 아이패드
IPAD_MODELS_HOT = ['Pro M4 13', 'Pro M4 11', 'Air M2 13', 'Air M2 11', 'Pro M2 12.9', 'Pro M2 11']
IPAD_MODELS_MID = ['Air 5', 'Air 4', 'mini 7', 'mini 6', '10세대', '9세대', 'Pro 12.9 5', 'Pro 11 3']
IPAD_MODELS_LOW = ['8세대', '7세대', 'Air 3', 'mini 5']

IPAD_SCREEN_HOT = ['액정 수리', '액정 교체', '액정 깨짐', '화면 깨짐', '디스플레이 교체']
IPAD_SCREEN_MID = ['액정', '화면 수리', '화면 교체', '터치 고장', '터치 안됨',
                   '액정 줄', '화면 멍', '유리 깨짐', '액정 수리비', '액정 가격',
                   '화면 먹통', '액정 갈라짐', '액정 터치', '디지타이저 교체']
IPAD_SCREEN_LOW = ['액정 당일', '액정 전문', '액정 저렴', '액정 긴급', '액정 출장',
                   '액정 할인', '액정 견적', '액정 보증', '액정 잘하는곳', '액정 후기']

IPAD_OTHER_TYPES = ['배터리 교체', '충전 포트', '카메라 수리', '침수 복구',
                    '홈버튼 수리', '프레임 수리', '스피커 수리']

# 애플워치
WATCH_SERIES_HOT = ['Series 10', 'Series 9', 'Ultra 2', 'Ultra', 'SE 2']
WATCH_SERIES_MID = ['Series 8', 'Series 7', 'Series 6', 'Series 5', 'SE']
WATCH_SERIES_LOW = ['Series 4', 'Series 3']

WATCH_SCREEN_HOT = ['액정 교체', '액정 수리', '액정 깨짐', '화면 깨짐', '디스플레이 교체']
WATCH_SCREEN_MID = ['액정', '화면 수리', '유리 깨짐', '전면유리 교체', '액정 줄',
                    '화면 초록줄', '액정 수리비', '액정 가격', '액정 후기', '터치 안됨']
WATCH_SCREEN_LOW = ['액정 당일', '액정 전문', '액정 저렴', '액정 할인', '액정 보증',
                    '액정 견적', '액정 긴급', '액정 정품', '액정 잘하는곳']

WATCH_OTHER_TYPES = ['배터리 교체', '후면유리', '크라운 수리', '침수 복구', '버튼 수리', '전원 수리']

# 애플펜슬 (액정 없음 — 기존 방식 유지)
PENCIL_MODELS = ['2세대', '1세대', 'Pro', 'USB-C']
PENCIL_SYMPTOMS_HOT = ['수리', '배터리 교체', '충전 안됨', '연결 안됨']
PENCIL_SYMPTOMS_MID = ['인식 안됨', '필기 안됨', '끊김', '지연', '리퍼', '교체', '팁 교체',
                       '가격', '수리 비용', '공식 AS', '사설 수리', '중고']
PENCIL_SYMPTOMS_LOW = ['고장', '오작동', '무상 교체', '정품 확인', '방문수리', '출장수리', '할인']

REGIONS = {
    '가산': ['가산', '가산디지털단지', '가산동', '독산동', '시흥동', '금천구', '가리봉동',
             '구로디지털단지', '구로동', '대림동'],
    '신림': ['신림', '신림동', '봉천동', '서울대입구', '신대방', '관악구', '사당',
             '남현동', '상도동', '대방동'],
    '목동': ['목동', '목동역', '양천구', '신정동', '등촌동', '화곡동', '신월동',
             '영등포', '당산', '여의도']
}


# ═════════════════════════════════════════════════
#  키워드 생성 — 액정 중심 (70%)
# ═════════════════════════════════════════════════

def gen_iphone_general():
    keywords = []

    # 【액정 70개】
    # 1-1) 단독 + 액정 표현 (최우선)
    for term in SCREEN_TERMS_HOT:
        keywords.append({'keyword': f'아이폰 {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-메인'})

    # 1-2) 최신 모델 × 주요 액정 표현
    for model in IPHONE_MODELS_HOT:
        for term in SCREEN_TERMS_HOT:
            keywords.append({'keyword': f'아이폰 {model} {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-최신모델'})

    # 1-3) 최신 모델 × 기타 액정 표현
    for model in IPHONE_MODELS_HOT[:4]:
        for term in SCREEN_TERMS_MID[:4]:
            keywords.append({'keyword': f'아이폰 {model} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-최신세부'})

    # 1-4) 중간 모델 × 핵심 액정
    for model in IPHONE_MODELS_MID[:6]:
        for term in SCREEN_TERMS_HOT[:3]:
            keywords.append({'keyword': f'아이폰 {model} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-중간'})

    # 1-5) 구형 + 액정
    for model in IPHONE_MODELS_LOW[:5]:
        keywords.append({'keyword': f'아이폰 {model} 액정 수리', 'tier': 2, 'cat': 'screen', 'group': '액정-구형'})

    # 1-6) 액정 기타 변형
    for term in SCREEN_TERMS_LOW[:10]:
        keywords.append({'keyword': f'아이폰 {term}', 'tier': 2, 'cat': 'screen', 'group': '액정-롱테일'})

    # 【기타 수리 30개】
    for term in IPHONE_OTHER_TYPES:
        keywords.append({'keyword': f'아이폰 {term}', 'tier': 1, 'cat': 'other', 'group': '기타-메인'})

    for model in IPHONE_MODELS_HOT[:5]:
        for term in IPHONE_OTHER_TYPES[:3]:
            keywords.append({'keyword': f'아이폰 {model} {term}', 'tier': 2, 'cat': 'other', 'group': '기타-모델'})

    # 추가 일반
    keywords.extend([
        {'keyword': '아이폰 수리', 'tier': 0, 'cat': 'other', 'group': '기타-메인'},
        {'keyword': '아이폰 수리비', 'tier': 1, 'cat': 'other', 'group': '기타-비용'},
        {'keyword': '아이폰 당일수리', 'tier': 1, 'cat': 'other', 'group': '기타-당일'},
        {'keyword': '아이폰 수리 잘하는곳', 'tier': 2, 'cat': 'other', 'group': '기타-검색'},
        {'keyword': '아이폰 전문 수리점', 'tier': 2, 'cat': 'other', 'group': '기타-검색'},
    ])

    return _dedupe_and_trim(keywords, 100)


def gen_ipad_general():
    keywords = []

    # 액정 70개
    for term in IPAD_SCREEN_HOT:
        keywords.append({'keyword': f'아이패드 {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-메인'})

    for model in IPAD_MODELS_HOT:
        for term in IPAD_SCREEN_HOT:
            keywords.append({'keyword': f'아이패드 {model} {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-최신모델'})

    for model in IPAD_MODELS_HOT[:4]:
        for term in IPAD_SCREEN_MID[:5]:
            keywords.append({'keyword': f'아이패드 {model} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-세부'})

    for model in IPAD_MODELS_MID[:5]:
        for term in IPAD_SCREEN_HOT[:3]:
            keywords.append({'keyword': f'아이패드 {model} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-중간'})

    for term in IPAD_SCREEN_LOW[:10]:
        keywords.append({'keyword': f'아이패드 {term}', 'tier': 2, 'cat': 'screen', 'group': '액정-롱테일'})

    # 기타 30개
    for term in IPAD_OTHER_TYPES:
        keywords.append({'keyword': f'아이패드 {term}', 'tier': 1, 'cat': 'other', 'group': '기타-메인'})

    for model in IPAD_MODELS_HOT[:4]:
        for term in IPAD_OTHER_TYPES[:3]:
            keywords.append({'keyword': f'아이패드 {model} {term}', 'tier': 2, 'cat': 'other', 'group': '기타-모델'})

    keywords.extend([
        {'keyword': '아이패드 수리', 'tier': 0, 'cat': 'other', 'group': '기타-메인'},
        {'keyword': '아이패드 수리비', 'tier': 1, 'cat': 'other', 'group': '기타-비용'},
        {'keyword': '아이패드 당일수리', 'tier': 1, 'cat': 'other', 'group': '기타-당일'},
        {'keyword': '아이패드 전문점', 'tier': 2, 'cat': 'other', 'group': '기타-검색'},
    ])

    return _dedupe_and_trim(keywords, 100)


def gen_watch_general():
    keywords = []

    # 액정 65개
    for term in WATCH_SCREEN_HOT:
        keywords.append({'keyword': f'애플워치 {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-메인'})

    for series in WATCH_SERIES_HOT:
        for term in WATCH_SCREEN_HOT:
            keywords.append({'keyword': f'애플워치 {series} {term}', 'tier': 0, 'cat': 'screen', 'group': '액정-최신'})

    for series in WATCH_SERIES_HOT[:3]:
        for term in WATCH_SCREEN_MID[:5]:
            keywords.append({'keyword': f'애플워치 {series} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-세부'})

    for series in WATCH_SERIES_MID[:4]:
        for term in WATCH_SCREEN_HOT[:3]:
            keywords.append({'keyword': f'애플워치 {series} {term}', 'tier': 1, 'cat': 'screen', 'group': '액정-중간'})

    for term in WATCH_SCREEN_LOW[:8]:
        keywords.append({'keyword': f'애플워치 {term}', 'tier': 2, 'cat': 'screen', 'group': '액정-롱테일'})

    # 기타 35개
    for term in WATCH_OTHER_TYPES:
        keywords.append({'keyword': f'애플워치 {term}', 'tier': 1, 'cat': 'other', 'group': '기타-메인'})

    for series in WATCH_SERIES_HOT[:4]:
        for term in WATCH_OTHER_TYPES[:3]:
            keywords.append({'keyword': f'애플워치 {series} {term}', 'tier': 2, 'cat': 'other', 'group': '기타-모델'})

    for series in WATCH_SERIES_MID[:3]:
        keywords.append({'keyword': f'애플워치 {series} 배터리 교체', 'tier': 2, 'cat': 'other', 'group': '기타-배터리'})

    keywords.extend([
        {'keyword': '애플워치 수리', 'tier': 0, 'cat': 'other', 'group': '기타-메인'},
        {'keyword': 'Apple Watch 수리', 'tier': 1, 'cat': 'other', 'group': '기타-영문'},
        {'keyword': '애플워치 당일', 'tier': 1, 'cat': 'other', 'group': '기타-당일'},
    ])

    return _dedupe_and_trim(keywords, 100)


def gen_pencil_general():
    """애플펜슬은 액정 없으므로 기존 스타일 유지"""
    keywords = []
    for model in PENCIL_MODELS:
        for sym in PENCIL_SYMPTOMS_HOT:
            keywords.append({'keyword': f'애플펜슬 {model} {sym}', 'tier': 1, 'cat': 'other', 'group': '모델-핵심'})
        for sym in PENCIL_SYMPTOMS_MID:
            keywords.append({'keyword': f'애플펜슬 {model} {sym}', 'tier': 2, 'cat': 'other', 'group': '모델-세부'})

    for sym in PENCIL_SYMPTOMS_HOT:
        keywords.append({'keyword': f'애플펜슬 {sym}', 'tier': 0, 'cat': 'other', 'group': '메인'})
    for sym in PENCIL_SYMPTOMS_MID:
        keywords.append({'keyword': f'애플펜슬 {sym}', 'tier': 1, 'cat': 'other', 'group': '메인세부'})
    for sym in PENCIL_SYMPTOMS_LOW:
        keywords.append({'keyword': f'애플펜슬 {sym}', 'tier': 2, 'cat': 'other', 'group': '롱테일'})

    for sym in ['수리', '배터리 교체', '2세대 수리', '가격']:
        keywords.append({'keyword': f'Apple Pencil {sym}', 'tier': 1, 'cat': 'other', 'group': '영문'})

    keywords.extend([
        {'keyword': '애플펜슬 2세대 8만원', 'tier': 0, 'cat': 'other', 'group': '마케팅'},
        {'keyword': '애플펜슬 당일수리', 'tier': 0, 'cat': 'other', 'group': '마케팅'},
        {'keyword': '애플펜슬 저렴하게', 'tier': 1, 'cat': 'other', 'group': '마케팅'},
        {'keyword': '애플펜슬 수리점', 'tier': 1, 'cat': 'other', 'group': '마케팅'},
        {'keyword': '애플펜슬 방문수리', 'tier': 2, 'cat': 'other', 'group': '마케팅'},
    ])

    return _dedupe_and_trim(keywords, 100)


def gen_region_keywords(regions, device, screen_terms_hot, screen_terms_mid, other_terms, models_hot=None):
    """지역 광고그룹 — 액정 중심"""
    keywords = []

    # 【지역 × 액정 70개】
    for region in regions:
        for term in screen_terms_hot[:3]:
            tier = 0 if region in regions[:3] else 1
            keywords.append({'keyword': f'{region} {device} {term}', 'tier': tier, 'cat': 'screen', 'group': '지역-액정핵심'})

    # 지역 × 모델 × 액정 수리
    if models_hot:
        for region in regions[:5]:
            for model in models_hot[:4]:
                keywords.append({
                    'keyword': f'{region} {device} {model} 액정',
                    'tier': 1 if region in regions[:3] else 2,
                    'cat': 'screen',
                    'group': '지역-모델액정'
                })

    # 지역 × 기타 액정
    for region in regions[:5]:
        for term in screen_terms_mid[:4]:
            keywords.append({
                'keyword': f'{region} {device} {term}',
                'tier': 1 if region in regions[:2] else 2,
                'cat': 'screen',
                'group': '지역-액정세부'
            })

    # 【지역 × 기타 수리 30개】
    for region in regions:
        keywords.append({
            'keyword': f'{region} {device} 수리',
            'tier': 1 if region in regions[:3] else 2,
            'cat': 'other',
            'group': '지역-메인'
        })
        keywords.append({
            'keyword': f'{region} {device} 배터리',
            'tier': 2,
            'cat': 'other',
            'group': '지역-배터리'
        })

    # 기타 수리 × 지역 주요
    for region in regions[:4]:
        for term in other_terms[:3]:
            keywords.append({
                'keyword': f'{region} {device} {term}',
                'tier': 2,
                'cat': 'other',
                'group': '지역-부품'
            })

    # 지역 + 서비스 혜택
    for region in regions[:3]:
        keywords.append({'keyword': f'{region} {device} 당일', 'tier': 1, 'cat': 'other', 'group': '지역-당일'})
        keywords.append({'keyword': f'{region} {device} 수리비', 'tier': 2, 'cat': 'other', 'group': '지역-비용'})

    return _dedupe_and_trim(keywords, 100)


def gen_pencil_region(regions):
    keywords = []
    for region in regions:
        for sym in PENCIL_SYMPTOMS_HOT + PENCIL_SYMPTOMS_MID[:6]:
            keywords.append({
                'keyword': f'{region} 애플펜슬 {sym}',
                'tier': 1 if region in regions[:3] else 2,
                'cat': 'other',
                'group': '지역-증상'
            })
    for region in regions[:5]:
        for model in ['2세대', '1세대', 'Pro']:
            keywords.append({
                'keyword': f'{region} 애플펜슬 {model}',
                'tier': 2,
                'cat': 'other',
                'group': '지역-모델'
            })
    for region in regions[:3]:
        keywords.append({'keyword': f'{region} 애플펜슬 8만원', 'tier': 0, 'cat': 'other', 'group': '마케팅'})
    return _dedupe_and_trim(keywords, 100)


def _dedupe_and_trim(keywords, limit):
    seen = set()
    unique = []
    for k in keywords:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            unique.append(k)
    unique.sort(key=lambda x: (x['tier'], 0 if x['cat'] == 'screen' else 1, len(x['keyword'])))
    return unique[:limit]


# ═════════════════════════════════════════════════
#  입찰가 (2~3위 목표, 액정 +30% 할증)
# ═════════════════════════════════════════════════

def bid_general(device, tier, cat):
    base = {
        '아이폰':  {0: 2000, 1: 1200, 2: 600},
        '아이패드': {0: 1700, 1: 1000, 2: 550},
        '애플워치': {0: 1500, 1: 900, 2: 550},
        '애플펜슬': {0: 1000, 1: 700, 2: 450}
    }
    b = base[device][tier]
    if cat == 'screen':
        b = int(b * 1.3)  # 액정 +30% 할증
    return b


def bid_region(tier, cat):
    b = {0: 550, 1: 400, 2: 280}[tier]
    if cat == 'screen':
        b = int(b * 1.3)
    return b


# ═════════════════════════════════════════════════
#  엑셀 출력
# ═════════════════════════════════════════════════
HEADER_FILL = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')

TIER_FILL = {
    0: PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid'),
    1: PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
    2: PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
}
TIER_FONT = {
    0: Font(name='맑은 고딕', size=10, bold=True, color='991B1B'),
    1: Font(name='맑은 고딕', size=10, color='CA8A04'),
    2: Font(name='맑은 고딕', size=10, color='16A34A')
}
TIER_LABEL = {0: '★★프리미엄', 1: '★최우선', 2: '보조'}

SCREEN_FILL = PatternFill(start_color='FFF5F0', end_color='FFF5F0', fill_type='solid')  # 액정 행 강조

ZEBRA = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
BODY = Font(name='맑은 고딕', size=10)
THIN = Side(border_style='thin', color='E8E8E8')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


def add_sheet(wb, title, keywords, bid_fn, device_for_bid=None):
    ws = wb.create_sheet(title)
    headers = ['No', '키워드', '분류', '추천 입찰가(원)', '우선순위', '그룹']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    for idx, k in enumerate(keywords, start=1):
        bid = bid_fn(device_for_bid, k['tier'], k['cat']) if device_for_bid else bid_fn(k['tier'], k['cat'])
        cat_label = '🔴 액정' if k['cat'] == 'screen' else '⚪ 기타'
        ws.append([idx, k['keyword'], cat_label, bid, TIER_LABEL[k['tier']], k['group']])
        r = ws.max_row

        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = BORDER

        # 액정 행은 연한 오렌지 배경
        if k['cat'] == 'screen':
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = SCREEN_FILL

        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=4).alignment = RIGHT
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=5).alignment = CENTER
        ws.cell(row=r, column=5).fill = TIER_FILL[k['tier']]
        ws.cell(row=r, column=5).font = TIER_FONT[k['tier']]
        ws.cell(row=r, column=6).alignment = CENTER

    widths = [6, 35, 10, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    screen_count = sum(1 for k in keywords if k['cat'] == 'screen')
    other_count = len(keywords) - screen_count
    avg_bid = sum(bid_fn(device_for_bid, k['tier'], k['cat']) if device_for_bid else bid_fn(k['tier'], k['cat']) for k in keywords) // len(keywords)

    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1).value = '액정 키워드:'
    ws.cell(row=summary_row, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=summary_row, column=2).value = f'{screen_count}개 ({screen_count}%)'
    ws.cell(row=summary_row, column=2).font = Font(name='맑은 고딕', size=10, bold=True, color='E8732A')

    ws.cell(row=summary_row + 1, column=1).value = '기타 키워드:'
    ws.cell(row=summary_row + 1, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=summary_row + 1, column=2).value = f'{other_count}개'

    ws.cell(row=summary_row + 2, column=1).value = '평균 입찰가:'
    ws.cell(row=summary_row + 2, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=summary_row + 2, column=4).value = avg_bid
    ws.cell(row=summary_row + 2, column=4).number_format = '#,##0원'
    ws.cell(row=summary_row + 2, column=4).font = Font(name='맑은 고딕', size=10, bold=True, color='E8732A')
    ws.cell(row=summary_row + 2, column=4).alignment = RIGHT

    return {'screen_count': screen_count, 'other_count': other_count, 'avg_bid': avg_bid}


def add_summary_sheet(wb, data):
    ws = wb.create_sheet('0.전체 요약', 0)
    headers = ['캠페인', '광고그룹', '액정 키워드', '기타', '총 키워드', '평균 입찰가(원)', '비고']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    for row in data:
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = BORDER
            if (r - 1) % 2 == 1:
                cell.fill = ZEBRA
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=5).alignment = CENTER
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = '#,##0'

        # 액정 키워드 수 강조
        ws.cell(row=r, column=3).font = Font(name='맑은 고딕', size=10, bold=True, color='E8732A')

    widths = [16, 22, 12, 10, 12, 16, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 합계
    total_screen = sum(r[2] for r in data)
    total_other = sum(r[3] for r in data)
    total_kw = sum(r[4] for r in data)
    last = ws.max_row + 1
    ws.cell(row=last, column=1).value = '합계'
    ws.cell(row=last, column=3).value = total_screen
    ws.cell(row=last, column=4).value = total_other
    ws.cell(row=last, column=5).value = total_kw
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=last, column=c)
        cell.fill = PatternFill(start_color='E8732A', end_color='E8732A', fill_type='solid')
        cell.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        cell.border = BORDER
    ws.cell(row=last, column=3).alignment = CENTER
    ws.cell(row=last, column=4).alignment = CENTER
    ws.cell(row=last, column=5).alignment = CENTER

    # 액정 비율 계산
    ratio = total_screen * 100 // total_kw if total_kw else 0
    note_row = last + 2
    ws.cell(row=note_row, column=1).value = '💡 액정 비중:'
    ws.cell(row=note_row, column=1).font = Font(name='맑은 고딕', size=11, bold=True)
    ws.cell(row=note_row, column=2).value = f'{ratio}% (목표 70% 이상)'
    ws.cell(row=note_row, column=2).font = Font(name='맑은 고딕', size=11, bold=True, color='E8732A')

    ws.cell(row=note_row + 1, column=1).value = '💡 액정 키워드 입찰가는 기타 대비 +30% 할증'
    ws.cell(row=note_row + 1, column=1).font = Font(name='맑은 고딕', size=10, color='666666')

    ws.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
#  메인 실행
# ═════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

summary_data = []

# 아이폰 4그룹
iphone_kws = gen_iphone_general()
stats = add_sheet(wb, '1-A.아이폰 일반', iphone_kws, bid_general, '아이폰')
summary_data.append(['① 아이폰', '1-A 아이폰 일반', stats['screen_count'], stats['other_count'], len(iphone_kws), stats['avg_bid'], '메인 ★★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = gen_region_keywords(regions, '아이폰',
                              SCREEN_TERMS_HOT, SCREEN_TERMS_MID, IPHONE_OTHER_TYPES,
                              IPHONE_MODELS_HOT)
    label = f'1-{"BCD"[["가산","신림","목동"].index(region_label)]}.아이폰 {region_label}'
    stats = add_sheet(wb, label, kws, bid_region)
    summary_data.append(['① 아이폰', f'1-{"BCD"[["가산","신림","목동"].index(region_label)]} 아이폰 {region_label}', stats['screen_count'], stats['other_count'], len(kws), stats['avg_bid'], f'{region_label}점'])

# 아이패드 4그룹
ipad_kws = gen_ipad_general()
stats = add_sheet(wb, '2-A.아이패드 일반', ipad_kws, bid_general, '아이패드')
summary_data.append(['② 아이패드', '2-A 아이패드 일반', stats['screen_count'], stats['other_count'], len(ipad_kws), stats['avg_bid'], '보조 ★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = gen_region_keywords(regions, '아이패드',
                              IPAD_SCREEN_HOT, IPAD_SCREEN_MID, IPAD_OTHER_TYPES,
                              IPAD_MODELS_HOT)
    label = f'2-{"BCD"[["가산","신림","목동"].index(region_label)]}.아이패드 {region_label}'
    stats = add_sheet(wb, label, kws, bid_region)
    summary_data.append(['② 아이패드', f'2-{"BCD"[["가산","신림","목동"].index(region_label)]} 아이패드 {region_label}', stats['screen_count'], stats['other_count'], len(kws), stats['avg_bid'], f'{region_label}점'])

# 애플워치 4그룹
watch_kws = gen_watch_general()
stats = add_sheet(wb, '3-A.애플워치 일반', watch_kws, bid_general, '애플워치')
summary_data.append(['③ 애플워치', '3-A 애플워치 일반', stats['screen_count'], stats['other_count'], len(watch_kws), stats['avg_bid'], '보조 ★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = gen_region_keywords(regions, '애플워치',
                              WATCH_SCREEN_HOT, WATCH_SCREEN_MID, WATCH_OTHER_TYPES,
                              WATCH_SERIES_HOT)
    label = f'3-{"BCD"[["가산","신림","목동"].index(region_label)]}.애플워치 {region_label}'
    stats = add_sheet(wb, label, kws, bid_region)
    summary_data.append(['③ 애플워치', f'3-{"BCD"[["가산","신림","목동"].index(region_label)]} 애플워치 {region_label}', stats['screen_count'], stats['other_count'], len(kws), stats['avg_bid'], f'{region_label}점'])

# 애플펜슬 4그룹 (액정 없음)
pencil_kws = gen_pencil_general()
stats = add_sheet(wb, '4-A.애플펜슬 일반', pencil_kws, bid_general, '애플펜슬')
summary_data.append(['④ 애플펜슬', '4-A 애플펜슬 일반', stats['screen_count'], stats['other_count'], len(pencil_kws), stats['avg_bid'], '차별화 ★★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = gen_pencil_region(regions)
    label = f'4-{"BCD"[["가산","신림","목동"].index(region_label)]}.애플펜슬 {region_label}'
    stats = add_sheet(wb, label, kws, bid_region)
    summary_data.append(['④ 애플펜슬', f'4-{"BCD"[["가산","신림","목동"].index(region_label)]} 애플펜슬 {region_label}', stats['screen_count'], stats['other_count'], len(kws), stats['avg_bid'], f'{region_label}점'])

add_summary_sheet(wb, summary_data)
wb.save(OUT)

total_screen = sum(r[2] for r in summary_data)
total_kw = sum(r[4] for r in summary_data)
print('[OK] saved:', OUT)
print(f'  total groups: 16 / keywords: {total_kw} / screen: {total_screen} ({total_screen*100//total_kw}%)')
