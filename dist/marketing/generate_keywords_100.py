# -*- coding: utf-8 -*-
"""
다올리페어 네이버 파워링크 — 광고그룹당 100개 키워드 + 차등 입찰가
2~3위 노출 목표 기준 (2026년 4월 추정)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import itertools

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'naver-powerlink-keywords-100.xlsx')

# ═════════════════════════════════════════════════
#  키워드 풀 정의
# ═════════════════════════════════════════════════

# ── 아이폰 일반 ──
IPHONE_MODELS_HOT = ['17 Pro Max', '17 Pro', '17', '16 Pro Max', '16 Pro', '16']  # 최신 모델 (Tier 1)
IPHONE_MODELS_MID = ['15 Pro Max', '15 Pro', '15', '14 Pro Max', '14 Pro', '14', '13 Pro', '13', '12 Pro', '12', 'SE 3']  # 중간 (Tier 2)
IPHONE_MODELS_LOW = ['11 Pro Max', '11 Pro', '11', 'SE 2', 'XS', 'XR', 'X', '8']  # 구형 (Tier 3)

IPHONE_REPAIR_TYPES = ['수리', '액정 수리', '배터리 교체', '충전포트 수리', '카메라 수리',
                       '후면유리 교체', '침수 복구', '화면 깨짐', '버튼 수리', '스피커 수리']

# ── 아이패드 일반 ──
IPAD_MODELS_HOT = ['Pro M4 13', 'Pro M4 11', 'Air M2 13', 'Air M2 11', 'Pro M2', 'Air 5']
IPAD_MODELS_MID = ['Air 4', 'mini 7', 'mini 6', '10세대', '9세대', 'Pro 12.9 5', 'Pro 11 3']
IPAD_MODELS_LOW = ['8세대', '7세대', 'Air 3', 'mini 5']

IPAD_REPAIR_TYPES = ['수리', '액정 수리', '배터리 교체', '충전 포트', '카메라 수리',
                     '침수 복구', '화면 깨짐', '프레임 수리', '터치 오류', '홈버튼 수리']

# ── 애플워치 일반 ──
WATCH_SERIES_HOT = ['Series 10', 'Series 9', 'Ultra 2', 'Ultra', 'SE 2']
WATCH_SERIES_MID = ['Series 8', 'Series 7', 'Series 6', 'Series 5', 'SE']
WATCH_SERIES_LOW = ['Series 4', 'Series 3']

WATCH_REPAIR_TYPES = ['수리', '배터리 교체', '액정 교체', '후면유리', '크라운 수리',
                      '침수 복구', '전원 수리', '버튼 수리', '스피커 수리', '수압 테스트']

# ── 애플펜슬 일반 ──
PENCIL_MODELS = ['2세대', '1세대', 'Pro', 'USB-C']
PENCIL_SYMPTOMS = ['수리', '배터리 교체', '충전 안됨', '연결 안됨', '인식 안됨',
                   '필기 안됨', '끊김', '지연', '오작동', '고장',
                   '교체', '리퍼', '팁 교체', '공식 AS', '사설 수리',
                   '중고', '정품 확인', '가격', '수리 비용', '무상 교체']

# ── 지역 풀 ──
REGIONS = {
    '가산': ['가산', '가산디지털단지', '가산동', '독산동', '시흥동', '금천구', '가리봉동',
             '구로디지털단지', '구로동', '대림동', '남구로'],
    '신림': ['신림', '신림동', '봉천동', '서울대입구', '신대방', '관악구', '사당',
             '남현동', '상도동', '대방동'],
    '목동': ['목동', '목동역', '양천구', '신정동', '등촌동', '화곡동', '신월동',
             '영등포', '당산', '여의도', '강서구']
}

# 지역 × 기기별 조합 생성용
DEVICE_TYPES = ['아이폰', '아이패드', '애플워치', '애플펜슬']

REGION_REPAIR_SUFFIXES_IPHONE = ['수리', '액정', '액정 수리', '배터리', '배터리 교체',
                                  '충전포트', '침수', '후면유리', '카메라', '화면 깨짐',
                                  '17 수리', '17 액정', '16 수리', '16 액정', '16 배터리',
                                  '15 수리', '15 액정', '15 배터리', '14 수리', '14 배터리',
                                  '13 수리', '12 수리', '11 수리', 'SE 수리']

REGION_REPAIR_SUFFIXES_IPAD = ['수리', '액정 수리', '배터리 교체', '충전 포트', '침수 복구',
                               '화면 깨짐', '프로 수리', '에어 수리', '미니 수리',
                               '액정', '배터리', 'Pro 수리', 'Air 수리', 'mini 수리',
                               'M4 수리', 'M2 수리', '10세대 수리', '9세대 수리',
                               '프레임 수리', '터치 오류']

REGION_REPAIR_SUFFIXES_WATCH = ['수리', '배터리 교체', '액정 교체', '후면유리',
                                '크라운 수리', '침수', '배터리', '액정',
                                'Series 10 수리', 'Series 9 수리', 'Series 8 수리',
                                'Series 7 수리', 'Series 6 수리', 'Series 5 수리',
                                'Ultra 수리', 'Ultra 2 수리', 'SE 수리', 'SE 2 수리',
                                '전원 수리', '버튼 수리']

REGION_REPAIR_SUFFIXES_PENCIL = ['수리', '2세대 수리', '1세대 수리', 'Pro 수리',
                                  '배터리 교체', '충전 안됨', '연결 안됨', '인식 안됨',
                                  '필기 안됨', '리퍼', '교체', '팁 교체', '가격',
                                  '공식 AS', '사설 수리', '비용', '중고', '정품',
                                  '배터리', '고장']


# ═════════════════════════════════════════════════
#  키워드 생성 함수
# ═════════════════════════════════════════════════

def gen_general_keywords(device_label, hot, mid, low, types, base_keyword):
    """일반 광고그룹 키워드 생성 (100개 목표, Tier별 분류)"""
    keywords = []

    # 1) 단독 기기명 + 수리 타입 (메인 키워드)
    for t in types:
        keywords.append({
            'keyword': f'{device_label} {t}',
            'tier': 1,
            'group': '메인'
        })

    # 2) 최신 모델 × 수리 타입 (Tier 1)
    for model in hot:
        for t in types:
            keywords.append({
                'keyword': f'{device_label} {model} {t}',
                'tier': 1,
                'group': '최신모델'
            })

    # 3) 중간 모델 × 주요 수리 (Tier 2)
    for model in mid:
        for t in types[:5]:  # 상위 5개만
            keywords.append({
                'keyword': f'{device_label} {model} {t}',
                'tier': 2,
                'group': '중간모델'
            })

    # 4) 구형 모델 × 기본 수리 (Tier 3)
    for model in low:
        for t in types[:3]:  # 상위 3개만
            keywords.append({
                'keyword': f'{device_label} {model} {t}',
                'tier': 3,
                'group': '구형모델'
            })

    # 5) 기타 키워드 (정품/비용 관련)
    extra = [
        (f'{base_keyword} 비용', 2),
        (f'{base_keyword} 가격', 2),
        (f'{base_keyword} 후기', 2),
        (f'{base_keyword} 잘하는곳', 2),
        (f'{base_keyword} 당일', 1),
        (f'{base_keyword} 무료진단', 2),
        (f'{base_keyword} 전문점', 2),
        (f'서울 {base_keyword}', 2),
        (f'{base_keyword} 사설', 3),
        (f'{base_keyword} 공식 AS', 3),
    ]
    for kw, tier in extra:
        keywords.append({'keyword': kw, 'tier': tier, 'group': '기타'})

    # 100개만 유지 (Tier 1 우선)
    keywords.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    return keywords[:100]


def gen_pencil_general_keywords():
    """애플펜슬은 모델 적어서 다르게 생성 (100개 확장)"""
    keywords = []

    # 1) 기본 조합
    for model in PENCIL_MODELS:
        for sym in PENCIL_SYMPTOMS:
            keywords.append({
                'keyword': f'애플펜슬 {model} {sym}',
                'tier': 2 if sym in PENCIL_SYMPTOMS[:5] else 3,
                'group': '모델별'
            })

    # 2) 단독 애플펜슬 + 증상
    for sym in PENCIL_SYMPTOMS:
        keywords.append({
            'keyword': f'애플펜슬 {sym}',
            'tier': 1 if sym in ['수리', '배터리 교체', '충전 안됨', '연결 안됨'] else 2,
            'group': '메인증상'
        })

    # 3) Apple Pencil 영문
    for sym in ['수리', '배터리 교체', '2세대 수리', '리퍼', '가격', '고장']:
        keywords.append({
            'keyword': f'Apple Pencil {sym}',
            'tier': 2,
            'group': '영문'
        })

    # 4) 할인/특가 관련
    extra = [
        ('애플펜슬 2세대 8만원', 1), ('애플펜슬 저렴하게', 2),
        ('애플펜슬 당일수리', 1), ('애플펜슬 수리점', 2),
        ('애플펜슬 리퍼 가격', 2), ('애플펜슬 애플케어', 3),
        ('서울 애플펜슬 수리', 2), ('애플펜슬 수리 비용', 2),
        ('애플펜슬 방문수리', 2), ('애플펜슬 사설수리', 2),
    ]
    for kw, tier in extra:
        keywords.append({'keyword': kw, 'tier': tier, 'group': '마케팅'})

    # 중복 제거
    seen = set()
    unique = []
    for k in keywords:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            unique.append(k)

    unique.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    return unique[:100]


def gen_region_keywords(regions, device_prefixes):
    """지역 광고그룹 키워드 생성 (100개 목표)"""
    keywords = []

    # 1) 지역 × 기기 × 수리유형
    for region in regions:
        for device in device_prefixes:
            # 메인 키워드
            keywords.append({
                'keyword': f'{region} {device} 수리',
                'tier': 1 if region in regions[:3] else 2,
                'group': '메인'
            })
            keywords.append({
                'keyword': f'{region} {device} 액정',
                'tier': 1 if region in regions[:3] else 2,
                'group': '메인'
            })
            keywords.append({
                'keyword': f'{region} {device} 배터리',
                'tier': 2,
                'group': '부품'
            })
            keywords.append({
                'keyword': f'{region} {device} 수리비',
                'tier': 3,
                'group': '가격'
            })
            keywords.append({
                'keyword': f'{region} {device} 수리점',
                'tier': 3,
                'group': '위치'
            })

    # 2) 지역 + 모델
    common_models = {
        '아이폰': ['17', '16', '15', '14'],
        '아이패드': ['프로', '에어', '미니'],
        '애플워치': ['Ultra', 'SE'],
        '애플펜슬': ['2세대']
    }
    for region in regions[:5]:
        for device, models in common_models.items():
            for model in models:
                keywords.append({
                    'keyword': f'{region} {device} {model}',
                    'tier': 2,
                    'group': '모델별'
                })

    # 3) 중복 제거
    seen = set()
    unique = []
    for k in keywords:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            unique.append(k)

    unique.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    return unique[:100]


# ═════════════════════════════════════════════════
#  입찰가 결정 (2~3위 노출 목표, 2026년 4월 기준)
# ═════════════════════════════════════════════════

def bid_for_general(device, tier):
    """일반 광고그룹 입찰가"""
    base = {
        '아이폰': {1: 1800, 2: 1000, 3: 500},
        '아이패드': {1: 1500, 2: 900, 3: 500},
        '애플워치': {1: 1300, 2: 800, 3: 500},
        '애플펜슬': {1: 900, 2: 600, 3: 400}
    }
    return base[device][tier]


def bid_for_region(tier):
    """지역 광고그룹 입찰가 (지역이라 저렴)"""
    return {1: 500, 2: 350, 3: 250}[tier]


# ═════════════════════════════════════════════════
#  엑셀 스타일
# ═════════════════════════════════════════════════
HEADER_FILL = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
TIER_FILL = {
    1: PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
    2: PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
    3: PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
}
TIER_FONT = {
    1: Font(name='맑은 고딕', size=10, bold=True, color='DC2626'),
    2: Font(name='맑은 고딕', size=10, color='CA8A04'),
    3: Font(name='맑은 고딕', size=10, color='16A34A')
}
TIER_LABEL = {1: '★최우선', 2: '중간', 3: '보조'}
ZEBRA = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
BODY = Font(name='맑은 고딕', size=10)
THIN = Side(border_style='thin', color='E8E8E8')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


def add_sheet(wb, title, keywords, campaign, group, bid_fn, device_for_bid=None):
    """하나의 광고그룹 시트 추가"""
    ws = wb.create_sheet(title)
    headers = ['No', '키워드', '추천 입찰가(원)', '우선순위', '분류']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    # 캠페인·광고그룹 정보 상단 (row 2 공백, row 3 정보는 안 넣고 A열 메타로)

    for idx, k in enumerate(keywords, start=1):
        if device_for_bid:
            bid = bid_fn(device_for_bid, k['tier'])
        else:
            bid = bid_fn(k['tier'])

        ws.append([idx, k['keyword'], bid, TIER_LABEL[k['tier']], k['group']])
        r = ws.max_row

        # 스타일
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = BORDER
            if (r - 2) % 2 == 1:
                cell.fill = ZEBRA

        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=4).fill = TIER_FILL[k['tier']]
        ws.cell(row=r, column=4).font = TIER_FONT[k['tier']]
        ws.cell(row=r, column=5).alignment = CENTER

    # 컬럼 너비
    widths = [6, 35, 14, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    # 총계 행
    total_bid = sum(bid_fn(device_for_bid, k['tier']) if device_for_bid else bid_fn(k['tier']) for k in keywords)
    avg_bid = total_bid // len(keywords) if keywords else 0
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1).value = '평균 입찰가:'
    ws.cell(row=summary_row, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=summary_row, column=3).value = avg_bid
    ws.cell(row=summary_row, column=3).number_format = '#,##0원'
    ws.cell(row=summary_row, column=3).font = Font(name='맑은 고딕', size=10, bold=True, color='E8732A')
    ws.cell(row=summary_row, column=3).alignment = RIGHT
    ws.cell(row=summary_row + 1, column=1).value = '총 키워드:'
    ws.cell(row=summary_row + 1, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
    ws.cell(row=summary_row + 1, column=3).value = f'{len(keywords)}개'
    ws.cell(row=summary_row + 1, column=3).alignment = RIGHT


def add_summary_sheet(wb, data):
    """전체 요약 시트"""
    ws = wb.create_sheet('0.전체 요약', 0)
    headers = ['캠페인', '광고그룹', '키워드 수', '평균 입찰가(원)', '권장 일 예산(원)', '비고']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    for row in data:
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = BORDER
            if (r - 1) % 2 == 1:
                cell.fill = ZEBRA
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=4).alignment = RIGHT
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).number_format = '#,##0'

    widths = [18, 25, 12, 16, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 합계
    total_kw = sum(r[2] for r in data)
    total_budget = sum(r[4] for r in data)
    last = ws.max_row + 1
    ws.cell(row=last, column=1).value = '합계'
    ws.cell(row=last, column=3).value = total_kw
    ws.cell(row=last, column=5).value = total_budget
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=last, column=c)
        cell.fill = PatternFill(start_color='E8732A', end_color='E8732A', fill_type='solid')
        cell.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        cell.border = BORDER
    ws.cell(row=last, column=3).alignment = RIGHT
    ws.cell(row=last, column=5).alignment = RIGHT
    ws.cell(row=last, column=5).number_format = '#,##0'

    ws.freeze_panes = 'A2'


# ═════════════════════════════════════════════════
#  메인
# ═════════════════════════════════════════════════
wb = Workbook()
# 기본 시트 제거
wb.remove(wb.active)

summary_data = []

# ── 아이폰 4개 광고그룹 ──
iphone_general = gen_general_keywords('아이폰', IPHONE_MODELS_HOT, IPHONE_MODELS_MID, IPHONE_MODELS_LOW, IPHONE_REPAIR_TYPES, '아이폰 수리')
add_sheet(wb, '1-A.아이폰 일반', iphone_general, '아이폰 수리', '일반', bid_for_general, '아이폰')
avg = sum(bid_for_general('아이폰', k['tier']) for k in iphone_general) // len(iphone_general)
summary_data.append(['① 아이폰 수리', '1-A 아이폰 일반', len(iphone_general), avg, avg * len(iphone_general) // 10, '메인 매출 ★★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = gen_region_keywords(regions, ['아이폰'])
    # 광고그룹당 100개 목표. 아이폰 단독으로 50개 정도 나오므로 아이폰 모델 확장
    kws_expanded = list(kws)
    for region in regions:
        for suffix in REGION_REPAIR_SUFFIXES_IPHONE:
            kws_expanded.append({
                'keyword': f'{region} 아이폰 {suffix}',
                'tier': 2 if region in regions[:3] else 3,
                'group': '모델·부품'
            })
    # 중복 제거
    seen = set()
    final = []
    for k in kws_expanded:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            final.append(k)
    final.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    final = final[:100]

    label = f'1-{"BCD"[["가산","신림","목동"].index(region_label)]}.아이폰 {region_label}'
    add_sheet(wb, label, final, '아이폰 수리', region_label, bid_for_region)
    avg = sum(bid_for_region(k['tier']) for k in final) // len(final)
    summary_data.append(['① 아이폰 수리', f'1-{"BCD"[["가산","신림","목동"].index(region_label)]} 아이폰 {region_label}', len(final), avg, avg * len(final) // 10, f'{region_label}점 지역'])

# ── 아이패드 ──
ipad_general = gen_general_keywords('아이패드', IPAD_MODELS_HOT, IPAD_MODELS_MID, IPAD_MODELS_LOW, IPAD_REPAIR_TYPES, '아이패드 수리')
add_sheet(wb, '2-A.아이패드 일반', ipad_general, '아이패드 수리', '일반', bid_for_general, '아이패드')
avg = sum(bid_for_general('아이패드', k['tier']) for k in ipad_general) // len(ipad_general)
summary_data.append(['② 아이패드 수리', '2-A 아이패드 일반', len(ipad_general), avg, avg * len(ipad_general) // 10, '보조 매출 ★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = []
    for region in regions:
        for suffix in REGION_REPAIR_SUFFIXES_IPAD:
            kws.append({
                'keyword': f'{region} 아이패드 {suffix}',
                'tier': 1 if region == regions[0] else (2 if region in regions[:4] else 3),
                'group': '지역·부품'
            })
    # 기본 지역+단독
    for region in regions:
        kws.insert(0, {'keyword': f'{region} 아이패드 수리', 'tier': 1 if region in regions[:3] else 2, 'group': '메인'})

    seen = set()
    final = []
    for k in kws:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            final.append(k)
    final.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    final = final[:100]

    label = f'2-{"BCD"[["가산","신림","목동"].index(region_label)]}.아이패드 {region_label}'
    add_sheet(wb, label, final, '아이패드 수리', region_label, bid_for_region)
    avg = sum(bid_for_region(k['tier']) for k in final) // len(final)
    summary_data.append(['② 아이패드 수리', f'2-{"BCD"[["가산","신림","목동"].index(region_label)]} 아이패드 {region_label}', len(final), avg, avg * len(final) // 10, f'{region_label}점 지역'])

# ── 애플워치 ──
watch_general = gen_general_keywords('애플워치', WATCH_SERIES_HOT, WATCH_SERIES_MID, WATCH_SERIES_LOW, WATCH_REPAIR_TYPES, '애플워치 수리')
add_sheet(wb, '3-A.애플워치 일반', watch_general, '애플워치 수리', '일반', bid_for_general, '애플워치')
avg = sum(bid_for_general('애플워치', k['tier']) for k in watch_general) // len(watch_general)
summary_data.append(['③ 애플워치 수리', '3-A 애플워치 일반', len(watch_general), avg, avg * len(watch_general) // 10, '보조 매출 ★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = []
    for region in regions:
        for suffix in REGION_REPAIR_SUFFIXES_WATCH:
            kws.append({
                'keyword': f'{region} 애플워치 {suffix}',
                'tier': 1 if region == regions[0] else (2 if region in regions[:4] else 3),
                'group': '지역·부품'
            })

    seen = set()
    final = []
    for k in kws:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            final.append(k)
    final.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    final = final[:100]

    label = f'3-{"BCD"[["가산","신림","목동"].index(region_label)]}.애플워치 {region_label}'
    add_sheet(wb, label, final, '애플워치 수리', region_label, bid_for_region)
    avg = sum(bid_for_region(k['tier']) for k in final) // len(final)
    summary_data.append(['③ 애플워치 수리', f'3-{"BCD"[["가산","신림","목동"].index(region_label)]} 애플워치 {region_label}', len(final), avg, avg * len(final) // 10, f'{region_label}점 지역'])

# ── 애플펜슬 ──
pencil_general = gen_pencil_general_keywords()
add_sheet(wb, '4-A.애플펜슬 일반', pencil_general, '애플펜슬 수리', '일반', bid_for_general, '애플펜슬')
avg = sum(bid_for_general('애플펜슬', k['tier']) for k in pencil_general) // len(pencil_general)
summary_data.append(['④ 애플펜슬 수리', '4-A 애플펜슬 일반', len(pencil_general), avg, avg * len(pencil_general) // 10, '차별화 무기 ★★★'])

for region_label, regions in [('가산', REGIONS['가산']), ('신림', REGIONS['신림']), ('목동', REGIONS['목동'])]:
    kws = []
    for region in regions:
        for suffix in REGION_REPAIR_SUFFIXES_PENCIL:
            kws.append({
                'keyword': f'{region} 애플펜슬 {suffix}',
                'tier': 1 if region == regions[0] else (2 if region in regions[:4] else 3),
                'group': '지역·증상'
            })

    seen = set()
    final = []
    for k in kws:
        if k['keyword'] not in seen:
            seen.add(k['keyword'])
            final.append(k)
    final.sort(key=lambda x: (x['tier'], len(x['keyword'])))
    final = final[:100]

    label = f'4-{"BCD"[["가산","신림","목동"].index(region_label)]}.애플펜슬 {region_label}'
    add_sheet(wb, label, final, '애플펜슬 수리', region_label, bid_for_region)
    avg = sum(bid_for_region(k['tier']) for k in final) // len(final)
    summary_data.append(['④ 애플펜슬 수리', f'4-{"BCD"[["가산","신림","목동"].index(region_label)]} 애플펜슬 {region_label}', len(final), avg, avg * len(final) // 10, f'{region_label}점 지역'])

# 요약 시트를 맨 앞에 추가
add_summary_sheet(wb, summary_data)

wb.save(OUT)
print('[OK] saved:', OUT)
print('  total groups: 16 / keywords per group: ~100 / total:', sum(r[2] for r in summary_data))
